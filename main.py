import pandas as pd
import matplotlib.pyplot as plt
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font

root=Tk()
root.withdraw()

csv_path=filedialog.askopenfilename(filetypes=[("CSV Files","*.csv")])
if not csv_path:
    print("No file selected.")
    raise SystemExit

df=pd.read_csv(csv_path)

summary=df.describe(include="all")
pivot=df.select_dtypes(include="number").sum().to_frame("Total")

plt.figure(figsize=(6,4))
pivot.plot(kind="bar", legend=False)
plt.tight_layout()
chart_path="chart.png"
plt.savefig(chart_path)
plt.close()

out="report.xlsx"
with pd.ExcelWriter(out, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Data", index=False)
    summary.to_excel(writer, sheet_name="Summary")
    pivot.to_excel(writer, sheet_name="Totals")

wb=load_workbook(out)
ws=wb["Data"]
for c in ws[1]:
    c.font=Font(bold=True)
wb.save(out)
print("Report generated:", out)
